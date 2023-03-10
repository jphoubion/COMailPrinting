import json

from PySide6.QtWidgets import QFileDialog
from openpyxl import Workbook, load_workbook
import sqlite3

import re

import sqlmanagement
# TODO: Corriger l'import des personnes avec noms composés, plus de tiret etc

class ImportFromXls:

    def __init__(self):
        self.wb = Workbook()
        self.sheet = None

    def select_customers_file(self):
        dlg = QFileDialog(filter="*.xlsx;;*.xls")
        dlg.setWindowTitle("Selectionner le fichier des clients")
        dlg.setFileMode(QFileDialog.ExistingFile)
        dlg.setViewMode(QFileDialog.Detail)
        if dlg.exec():
            self.filename = dlg.selectedFiles()
            self.wb = load_workbook(self.filename[0])
            return self.wb.active

    def import_customers(self, conn, db_cursor):
        self.sheet = self.select_customers_file()

        # Fill CO table first
        ######################
        self.import_co(conn, db_cursor)

        db_connection = conn
        if self.sheet is not None:
            for row in self.sheet.iter_rows(min_row=2):
                co_name = str(row[3].value).upper().replace("'","").replace("C/O ME ",'').replace("C/O ME. ",'').\
                    replace("C/O MAITRE ",'').replace('MAITRE ','').replace('MAÎTRE ','').lstrip()
                co_id = sqlmanagement.get_result(db_connection, f"SELECT * FROM co WHERE co_name='{co_name}'")
                # remove weird characters from the name
                customer = str(row[2].value).replace("(", "").replace(")", "").replace("*","").replace("!","").replace("'", " ")
                req_customers = "INSERT INTO customers (customer_code, customer_name, co_id) VALUES "
                req_customers += f"('{row[0].value}', '{customer}', '{co_id[0][0]}')"
                print(req_customers)
                try:
                    db_cursor.execute(req_customers)
                    db_connection.commit()
                except sqlite3.IntegrityError as e:
                    print(e)
                    print(req_customers)


    def import_co(self, conn, db_cursor):
        db_connection = conn
        if self.sheet is not None:
            for row in self.sheet.iter_rows(min_row=2):
                req_co = "INSERT INTO co (co_name, co_address, is_lawyer, is_cpas) VALUES "
                co_name = str(row[3].value).upper().replace("'","").replace("C/O ME ",'').\
                    replace("C/O ME. ",'').replace("C/O MAITRE ",'').replace('MAITRE ','').replace('MAÎTRE ','').lstrip()
                address = f"{row[5].value}\n{row[7].value}\n{row[8].value}".replace("'", ' ')
                co_code = ["C/O ME", "C/O ME.", "CO MAITRE", "MAITRE", "MAÎTRE"]
                if any([code in str(row[3].value).upper() for code in co_code]):
                    lawyer = 1
                else:
                    lawyer = 0
                if "CPAS" in str(row[3].value).upper():
                    cpas = 1
                else:
                    cpas = 0
                req_co += f"('{co_name.upper()}', '{address.upper()}', '{lawyer}', '{cpas}'),"
                try:
                    db_cursor.execute(req_co[:-1])
                    db_connection.commit()
                except sqlite3.IntegrityError as e:
                    print(e)
                    # print(req_co)

    def clean_string(self, string):
        # remove weird characters from the name but keeps hyphen
        pattern = re.compile("[\W_0-9]+")
        dirty_name = str(string).replace("'", " ").split()
        cleaned_list = [pattern.sub('', word) for word in dirty_name]
        return ' '.join(cleaned_list).upper()